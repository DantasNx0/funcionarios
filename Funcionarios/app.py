import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os
from cadastro_page import CadastroFuncionarioPage
from listagem_page import ListagemFuncionariosPage

class MercadinhoDoZeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mercadinho do Sr. Zé ツ | Funcionários ")

        # Definindo o estilo da interface gráfica
        self.style = ttk.Style()
        self.style.configure('TFrame', background='#B3CCFF')
        self.style.configure('TButton', background='#B3CCFF')
        self.style.configure('TLabel', background='#E6F2FF', font=('Arial', 12))
        self.style.configure('Header.TLabel', background='#5C9BD0', foreground='white', font=('Arial', 25, 'bold'))
        self.style.configure('TEntry', width=30, font=('Arial', 12))
        self.style.configure('TCombobox', width=21, font=('Arial', 12))

        # Criando o arquivo de dados se não existir
        self.filename = "BD/funcionarios.xlsx"
        self.create_file_if_not_exists()

        # Variável para armazenar o próximo ID disponível
        self.next_id = self.get_next_id()

        # Configuração da página principal
        self.setup_main_page()

    def create_file_if_not_exists(self):
        """Cria o arquivo de dados se ele não existir."""
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "Nome", "CPF", "Gênero", "Cargo"])
            wb.save(self.filename)

    def setup_main_page(self):
        """Configura a página principal da aplicação."""
        # Frame principal
        self.main_frame = ttk.Frame(self.root, style='TFrame')
        self.main_frame.grid(row=0, column=0, padx=150, pady=200)

        # Cabeçalho
        ttk.Label(self.main_frame, text="Mercadinho do Sr. Zé ツ ", style='Header.TLabel').grid(row=0, column=0, columnspan=2, pady=10)

        # Botões principais
        ttk.Button(self.main_frame, text="Cadastrar Funcionário", command=self.show_cadastro_page).grid(row=1, column=0, padx=10, pady=10, sticky='ew')
        ttk.Button(self.main_frame, text="Listar Funcionários", command=self.show_listagem_page).grid(row=1, column=1, padx=10, pady=10, sticky='ew')

    def show_cadastro_page(self):
        """Mostra a página de cadastro de funcionários."""
        self.main_frame.grid_forget()
        CadastroFuncionarioPage(self.root, self)

    def show_listagem_page(self):
        """Mostra a página de listagem de funcionários."""
        self.main_frame.grid_forget()
        ListagemFuncionariosPage(self.root, self)

    def get_next_id(self):
        """Retorna o próximo ID disponível para cadastro."""
        wb = load_workbook(self.filename)
        ws = wb.active

        # Obter todos os IDs existentes
        ids = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0] is not None]

        # Verificar se não há IDs
        if not ids:
            return 1

        # Retornar o próximo ID disponível
        return max(ids) + 1

    def clear_entries(self, *entries):
        """Limpa os campos de entrada passados como argumento."""
        for entry in entries:
            entry.delete(0, tk.END)

    def validate_input(self, nome, cpf, genero, cargo):
        """Valida os dados inseridos no cadastro de funcionários."""
        if not nome or not cpf or not genero or not cargo:
            messagebox.showwarning("Atenção", "Todos os campos são obrigatórios")
            return False
        return True

    def cpf_exists(self, cpf, current_id=None):
        """Verifica se o CPF já está cadastrado para outro funcionário."""
        wb = load_workbook(self.filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == cpf:
                if current_id is None or current_id == '' or int(row[0]) != int(current_id):
                    return True
        return False

def main():
    root = tk.Tk()
    app = MercadinhoDoZeApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
