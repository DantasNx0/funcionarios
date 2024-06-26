from tkinter import messagebox, ttk
from openpyxl import load_workbook

class EdicaoFuncionarioPage:
    def __init__(self, root, app, cpf):
        self.root = root  # Armazena a referência para a janela principal
        self.app = app    # Armazena a referência para a aplicação principal (MercadinhoDoZeApp)
        self.cpf = cpf    # Armazena o CPF do funcionário a ser editado

        # Cria um frame para a interface gráfica
        self.frame = ttk.Frame(self.root, style='TFrame')
        self.frame.grid(row=0, column=0, padx=20, pady=20)

        # Título da página de edição de funcionário
        ttk.Label(self.frame, text="Edição de Funcionário", style='Header.TLabel').grid(row=0, column=0, columnspan=2, pady=10)

        # Configuração do formulário de edição
        self.setup_form()

        # Preenche os campos com os dados do funcionário pelo CPF
        self.preencher_dados()

    def setup_form(self):
        # Labels e campos de entrada para CPF, nome, gênero e cargo
        ttk.Label(self.frame, text="CPF:", style='TLabel').grid(row=1, column=0, padx=10, pady=10)
        self.entry_cpf = ttk.Entry(self.frame, style='TEntry', state='readonly')
        self.entry_cpf.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(self.frame, text="Nome:", style='TLabel').grid(row=2, column=0, padx=10, pady=10)
        self.entry_nome = ttk.Entry(self.frame, style='TEntry')
        self.entry_nome.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(self.frame, text="Gênero:", style='TLabel').grid(row=3, column=0, padx=10, pady=10)
        self.combo_genero = ttk.Combobox(self.frame, values=["Masculino", "Feminino", "Não identificado"], style='TCombobox')
        self.combo_genero.grid(row=3, column=1, padx=10, pady=10)

        ttk.Label(self.frame, text="Cargo:", style='TLabel').grid(row=4, column=0, padx=10, pady=10)
        self.entry_cargo = ttk.Entry(self.frame, style='TEntry')
        self.entry_cargo.grid(row=4, column=1, padx=10, pady=10)

        # Botões para salvar alterações e cancelar
        ttk.Button(self.frame, text="Salvar Alterações", command=self.salvar).grid(row=5, column=0, columnspan=2, padx=10, pady=10, sticky='ew')
        ttk.Button(self.frame, text="Cancelar", command=self.cancelar).grid(row=6, column=0, columnspan=2, padx=10, pady=10, sticky='ew')

    def preencher_dados(self):
        # Preenche os campos de entrada com os dados do funcionário baseado no CPF
        wb = load_workbook(self.app.filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[2].value == self.cpf:
                self.entry_cpf.insert(0, row[2].value)
                self.entry_nome.insert(0, row[1].value)
                self.combo_genero.set(row[3].value)
                self.entry_cargo.insert(0, row[4].value)
                break

    def salvar(self):
        # Método para salvar as alterações feitas no funcionário
        nome = self.entry_nome.get()
        genero = self.combo_genero.get()
        cargo = self.entry_cargo.get()

        # Valida os campos de entrada
        if not self.app.validate_input(nome, self.cpf, genero, cargo):
            return

        # Abre o arquivo Excel, atualiza as informações do funcionário com base no CPF
        wb = load_workbook(self.app.filename)
        ws = wb.active

        found = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[2].value == self.cpf:
                row[1].value = nome
                row[3].value = genero
                row[4].value = cargo
                wb.save(self.app.filename)
                messagebox.showinfo("Sucesso", "Funcionário editado com sucesso")
                found = True
                break

        # Exibe mensagem de erro se o CPF não for encontrado
        if not found:
            messagebox.showwarning("Erro", f"Funcionário com CPF {self.cpf} não encontrado.")

        # Volta para a página principal após salvar ou cancelar
        self.frame.grid_forget()
        self.app.setup_main_page()

    def cancelar(self):
        # Método para cancelar a edição e voltar para a página principal
        self.frame.grid_forget()
        self.app.setup_main_page()
