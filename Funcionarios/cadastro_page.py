import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook

class CadastroFuncionarioPage:
    def __init__(self, root, app):
        self.root = root
        self.app = app

        # Configuração do frame principal
        self.frame = ttk.Frame(self.root, style='TFrame')
        self.frame.grid(row=0, column=0, padx=150, pady=150)

        # Título da página de cadastro
        ttk.Label(self.frame, text="Cadastro de Funcionário", style='Header.TLabel').grid(row=0, column=0, columnspan=2, pady=10)

        # Configuração do formulário de cadastro
        self.setup_form()

    def setup_form(self):
        # Obtém o próximo ID disponível para cadastro
        self.next_id = self.app.next_id

        # Labels e campos de entrada para nome, CPF, gênero e cargo
        ttk.Label(self.frame, text="Nome:", style='TLabel').grid(row=1, column=0, padx=10, pady=10)
        self.entry_nome = ttk.Entry(self.frame, style='TEntry')
        self.entry_nome.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(self.frame, text="CPF:", style='TLabel').grid(row=2, column=0, padx=10, pady=10)
        self.entry_cpf = ttk.Entry(self.frame, style='TEntry')
        self.entry_cpf.grid(row=2, column=1, padx=10, pady=10)
        self.entry_cpf.config(validate='key', validatecommand=(self.root.register(self.validate_cpf), '%P'))

        ttk.Label(self.frame, text="Gênero:", style='TLabel').grid(row=3, column=0, padx=10, pady=10)
        self.combo_genero = ttk.Combobox(self.frame, values=["Masculino", "Feminino", "Não identificado"], style='TCombobox')
        self.combo_genero.grid(row=3, column=1, padx=10, pady=10)

        ttk.Label(self.frame, text="Cargo:", style='TLabel').grid(row=4, column=0, padx=10, pady=10)
        self.entry_cargo = ttk.Entry(self.frame, style='TEntry')
        self.entry_cargo.grid(row=4, column=1, padx=10, pady=10)

        # Botões para cadastrar e voltar
        ttk.Button(self.frame, text="Cadastrar", command=self.cadastrar).grid(row=5, column=0, columnspan=2, padx=10, pady=10, sticky='ew')
        ttk.Button(self.frame, text="Voltar", command=self.voltar).grid(row=6, column=0, columnspan=2, padx=10, pady=10, sticky='ew')

    def cadastrar(self):
        # Método para cadastrar um novo funcionário
        nome = self.entry_nome.get()
        cpf = self.entry_cpf.get()
        genero = self.combo_genero.get()
        cargo = self.entry_cargo.get()

        # Validação dos campos de entrada
        if not self.app.validate_input(nome, cpf, genero, cargo):
            return

        # Verifica se o CPF possui exatamente 11 dígitos numéricos
        if len(cpf) != 11 or not cpf.isdigit():
            messagebox.showwarning("Atenção", "CPF deve conter exatamente 11 dígitos numéricos")
            return

        # Verifica se o CPF já está cadastrado para outro funcionário
        if self.app.cpf_exists(cpf):
            messagebox.showwarning("Atenção", "CPF já cadastrado para outro funcionário")
            return

        # Abre o arquivo Excel, adiciona o novo funcionário e salva as alterações
        wb = load_workbook(self.app.filename)
        ws = wb.active
        next_row = ws.max_row + 1
        ws.append([self.next_id, nome, cpf, genero, cargo])
        wb.save(self.app.filename)

        # Exibe mensagem de sucesso, atualiza o próximo ID e limpa os campos de entrada
        messagebox.showinfo("Sucesso", "Funcionário cadastrado com sucesso")
        self.app.next_id += 1
        self.app.clear_entries(self.entry_nome, self.entry_cpf, self.entry_cargo)
        self.combo_genero.set('')
        self.entry_cpf.focus()

    def voltar(self):
        # Método para voltar à página principal
        self.frame.grid_forget()
        self.app.setup_main_page()

    def validate_cpf(self, cpf):
        # Função de validação para o campo de CPF permitindo apenas números e no máximo 11 dígitos
        if cpf.isdigit() and len(cpf) <= 11:
            return True , self.app.clear_entries(self.entry_cpf)# Validação passou e limpa o campo CPF
        else:
            return False  # Validação falhou

