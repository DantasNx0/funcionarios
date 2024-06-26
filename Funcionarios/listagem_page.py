# listagem_page.py
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from edicao_page import EdicaoFuncionarioPage  # Importando a classe EdicaoFuncionarioPage

class ListagemFuncionariosPage:
    def __init__(self, root, app):
        self.root = root          # Armazena a referência para a janela principal
        self.app = app            # Armazena a referência para a aplicação principal (MercadinhoDoZeApp)

        # Cria um frame para a interface gráfica
        self.frame = ttk.Frame(self.root, style='TFrame')
        self.frame.grid(row=0, column=0, padx=20, pady=20)

        # Título da página de listagem de funcionários
        ttk.Label(self.frame, text="Listagem de Funcionários", style='Header.TLabel').grid(row=0, column=0, columnspan=4, pady=10)

        # Configuração da listagem e botões
        self.setup_listagem()

    def setup_listagem(self):
        # Configuração da Treeview para exibir os funcionários
        self.tree = ttk.Treeview(self.frame, columns=("ID", "Nome", "CPF", "Gênero", "Cargo"), show="headings")
        self.tree.heading("ID", text="ID")
        self.tree.heading("Nome", text="Nome")
        self.tree.heading("CPF", text="CPF")
        self.tree.heading("Gênero", text="Gênero")
        self.tree.heading("Cargo", text="Cargo")
        self.tree.grid(row=1, column=0, columnspan=4, padx=10, pady=10)

        # Botões para atualizar lista, editar, excluir e voltar
        self.btn_atualizar = ttk.Button(self.frame, text="Atualizar Lista", command=self.atualizar_lista)
        self.btn_atualizar.grid(row=2, column=0, padx=10, pady=10, sticky='ew')

        self.btn_editar = ttk.Button(self.frame, text="Editar Selecionado", command=self.editar_selecionado)
        self.btn_editar.grid(row=2, column=1, padx=10, pady=10, sticky='ew')

        self.btn_excluir = ttk.Button(self.frame, text="Excluir Selecionado", command=self.confirmar_excluir_selecionado)
        self.btn_excluir.grid(row=2, column=2, padx=10, pady=10, sticky='ew')

        self.btn_voltar = ttk.Button(self.frame, text="Voltar para a Página Principal", command=self.voltar)
        self.btn_voltar.grid(row=2, column=3, padx=10, pady=10, sticky='ew')

        # Carrega os dados iniciais na Treeview
        self.load_data()

    def load_data(self):
        # Limpa os dados existentes na Treeview
        for row in self.tree.get_children():
            self.tree.delete(row)

        # Carrega os dados do arquivo Excel na Treeview
        wb = load_workbook(self.app.filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert('', tk.END, values=row)

    def atualizar_lista(self):
        # Atualiza os dados na Treeview
        self.load_data()

    def editar_selecionado(self):
        # Obtém o item selecionado na Treeview e abre a página de edição
        selected_item = self.tree.focus()
        if selected_item:
            data = self.tree.item(selected_item, 'values')
            cpf = data[2]  # CPF está na coluna de índice 2

            self.frame.grid_forget()
            EdicaoFuncionarioPage(self.root, self.app, cpf)

    def confirmar_excluir_selecionado(self):
        # Confirma a exclusão do funcionário selecionado na Treeview
        selected_item = self.tree.focus()
        if selected_item:
            data = self.tree.item(selected_item, 'values')
            cpf = data[2]  # CPF está na coluna de índice 2

            result = messagebox.askyesno("Confirmação", f"Tem certeza que deseja excluir o funcionário com CPF {cpf}?")
            if result:
                self.excluir_funcionario(cpf)

    def excluir_funcionario(self, cpf):
        # Exclui o funcionário do arquivo Excel
        wb = load_workbook(self.app.filename)
        ws = wb.active

        found = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[2].value == cpf:
                ws.delete_rows(row[0].row)
                wb.save(self.app.filename)
                messagebox.showinfo("Sucesso", f"Funcionário com CPF {cpf} excluído com sucesso")
                found = True
                break

        if not found:
            messagebox.showwarning("Erro", f"Funcionário com CPF {cpf} não encontrado")

        # Atualiza a lista após a exclusão
        self.load_data()

    def voltar(self):
        # Retorna para a página principal
        self.frame.grid_forget()
        self.app.setup_main_page()
